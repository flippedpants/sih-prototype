"""Generic mapping + CSV/Excel parser with per-row validation reporting."""
from __future__ import annotations

import csv
import io
import re
from collections.abc import Mapping
from typing import Any

from openpyxl import load_workbook
from rapidfuzz.fuzz import ratio

from .models import Entity, IngestionResult, Relationship, RowValidationError, SourceMapping
from .validation import parse_timestamp, parse_weight, require_value, resolve_mapping_columns, validate_structure


class SourceIngestionEngine:
    """Transform any config-compatible tabular file without source-specific code."""

    def ingest_bytes(self, file_name: str, content: bytes, mapping_data: SourceMapping | dict[str, Any]) -> IngestionResult:
        mapping = mapping_data if isinstance(mapping_data, SourceMapping) else SourceMapping.model_validate(mapping_data)
        headers, rows = _read_table(file_name, content)
        mapping = resolve_mapping_columns(headers, mapping, file_name)
        validate_structure(headers, rows, mapping)
        result = IngestionResult()
        entities: dict[str, Entity] = {}
        person_names: list[tuple[str, str]] = []
        for row_number, row in enumerate(rows, start=2):
            source_doc = f"{mapping.source_type}:{file_name}:{row_number}"
            row_person_names = person_names.copy()
            try:
                row_entities = [self._entity(block, row, source_doc, row_person_names) for block in mapping.entity_blocks()]
                row_relationships = [self._relationship(block, row, source_doc, mapping, row_person_names) for block in mapping.relationship_blocks()]
            except ValueError as error:
                result.validation_errors.append(RowValidationError(row_number=row_number, reason=str(error)))
                continue
            person_names = row_person_names
            for entity in row_entities:
                existing = entities.get(entity.id)
                if existing:
                    existing.aliases = list(dict.fromkeys(existing.aliases + entity.aliases))
                    existing.attributes.update(entity.attributes)
                    existing.source_docs = list(dict.fromkeys(existing.source_docs + entity.source_docs))
                else:
                    entities[entity.id] = entity
            result.relationships.extend(row_relationships)
        result.entities = list(entities.values())
        return result

    def _entity(self, block: Any, row: Mapping[str, object], source_doc: str, person_names: list[tuple[str, str]]) -> Entity:
        natural_key = require_value(row, block.id_column)
        canonical_name = _optional(row, block.canonical_name_column)
        entity_id = _entity_id(block.entity_type, natural_key, canonical_name, person_names)
        attributes = {column: value for column in block.attribute_columns if (value := _optional(row, column)) is not None}
        aliases = _aliases(_optional(row, block.aliases_column))
        return Entity(id=entity_id, type=block.entity_type, canonical_name=canonical_name, aliases=aliases, attributes=attributes, source_docs=[source_doc])

    def _relationship(self, block: Any, row: Mapping[str, object], source_doc: str, mapping: SourceMapping, person_names: list[tuple[str, str]]) -> Relationship:
        entity_by_column = {item.id_column: item for item in mapping.entity_blocks()}
        try:
            source_block, target_block = entity_by_column[block.source_column], entity_by_column[block.target_column]
        except KeyError as error:
            raise ValueError("relationship columns must reference configured entity id columns") from error
        source_id = _entity_id(source_block.entity_type, require_value(row, block.source_column), _optional(row, source_block.canonical_name_column), person_names)
        target_id = _entity_id(target_block.entity_type, require_value(row, block.target_column), _optional(row, target_block.canonical_name_column), person_names)
        return Relationship(source_id=source_id, target_id=target_id, type=block.type, weight=parse_weight(row, block.weight_columns), source_doc=source_doc, timestamp=parse_timestamp(row, block.timestamp_column))


def _read_table(file_name: str, content: bytes) -> tuple[list[str], list[dict[str, object]]]:
    suffix = file_name.lower().rsplit(".", 1)[-1] if "." in file_name else ""
    if suffix == "csv":
        try:
            reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        except UnicodeDecodeError as error:
            raise ValueError("CSV files must be UTF-8 encoded") from error
        return list(reader.fieldnames or []), list(reader)
    if suffix in {"xlsx", "xlsm"}:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(values, ())]
        return headers, [dict(zip(headers, row, strict=False)) for row in values]
    raise ValueError("supported source files are .csv, .xlsx, and .xlsm")


def _entity_id(entity_type: str, natural_key: str, canonical_name: str | None, person_names: list[tuple[str, str]]) -> str:
    if entity_type == "PERSON" and canonical_name:
        normalized_name = _normalize(canonical_name)
        for known_name, known_id in person_names:
            if normalized_name == known_name or ratio(normalized_name, known_name) >= 93:
                return known_id
        entity_id = f"PERSON:{_normalize(natural_key)}"
        person_names.append((normalized_name, entity_id))
        return entity_id
    return f"{entity_type}:{_normalize(natural_key)}"


def _optional(row: Mapping[str, object], column: str | None) -> str | None:
    if not column:
        return None
    value = row.get(column)
    text = str(value).strip() if value is not None else ""
    return text or None


def _aliases(value: str | None) -> list[str]:
    return [item.strip() for item in re.split(r"[|,]", value or "") if item.strip()]


def _normalize(value: str) -> str:
    return " ".join(value.strip().casefold().split())
